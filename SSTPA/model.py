import sys
import os
from gurobipy import Model, GRB, quicksum
import time
sys.path.append(os.path.abspath(os.path.join('..', 'ANFP-Calendar-Opt', 'SSTPA')))

from modules.params.params import N, F, S, I, T, G, R, EL, EV, L, RP, PI, EB, V, H, M, TIMELIMIT, START_TIME, stats, S_full
from modules.model_stats import ModelStats
import openpyxl as opx

m = Model("SSTPA MULTIPLES POSICIONES")


m.setParam('TimeLimit', TIMELIMIT)
m.setParam('MIPFocus', 1)
m.setParam('MIPGap', 0)


start_model = time.time()


#################
#*  VARIABLES  *#
#################


# x_nf: x[partido, fecha]
# 1 si el partido n se programa finalmente
# en la fecha f
# 0 en otro caso.
x = m.addVars(N, F, vtype=GRB.BINARY, name="x")

# y_is: y[equipo][patron_localias]
# 1 si al equipo i se le asigna el patron
# de localias s
# 0 en otro caso
y = {i: m.addVars(S[i], vtype=GRB.BINARY, name="y") for i in I}

#p_jilf: P[equipo, equipo, fecha, fecha]
#discreta, cant de puntos del equipo j al finalizar la fecha f
#con la info de los resultados hasta la fecha l inclusive
#en el MEJOR conjunto de resultados futuros para el equipo i
p_m = m.addVars(I,I,F,F, vtype=GRB.INTEGER, name="p_m")

#p_jilf: P[equipo, equipo, fecha, fecha]
#discreta, cant de puntos del equipo j al finalizar la fecha f
#con la info de los resultados hasta la fecha l inclusive
#en el PEOR conjunto de resultados futuros para el equipo i
p_p = m.addVars(I,I,F,F, vtype=GRB.INTEGER, name="p_p")


# v_nilf : v[partido, equipo, fecha, fecha]
# binaria,  1 si el equipo local gana el partido n
# de la fecha f teniendo informacion finalizada la fecha l
# en el MEJOR conjunto de resultados futuros para el equipo i
v_m = m.addVars(N,I,F,F, vtype=GRB.BINARY, name="v_m")

# v_nilf : v[partido, equipo, fecha, fecha]
# binaria,  1 si el equipo local gana el partido n
# de la fecha f teniendo informacion finalizada la fecha l
# en el MEJOR conjunto de resultados futuros para el equipo i
v_p = m.addVars(N,I,F,F, vtype=GRB.BINARY, name="v_p")


# a_nilf: a[partido,equipo,fecha,fecha]
# 1 si el equipo visitante gana el partido n de la fecha f
# teniendo información finalizada la fecha l
# en el MEJOR conjunto de resultados para el equipo i
a_m = m.addVars(N,I,F,F, vtype=GRB.BINARY, name="a_m")


# a_nilf: a[partido,equipo,fecha,fecha]
# 1 si el equipo visitante gana el partido n de la fecha f
# teniendo información finalizada la fecha l
# en el PEOR conjunto de resultados para el equipo i
a_p = m.addVars(N,I,F,F, vtype=GRB.BINARY, name="a_p")


#e_nilf: e[partido,equipo,fecha,fecha]
#binaria, toma el valor 1 si se empata el 
#partido n de la fecha f, con la info
#de los resultados hasta la fecha l inclusive
# en el MEJOR conjunto de resultados futuros para el euqipo i
e_m = m.addVars(N,I,F,F, vtype=GRB.BINARY, name="e_m")

#e_nilf: e[partido,equipo,fecha,fecha]
#binaria, toma el valor 1 si se empata el 
#partido n de la fecha f, con la info
#de los resultados hasta la fecha l inclusive
# en el PEOR- conjunto de resultados futuros para el euqipo i
e_p = m.addVars(N,I,F,F, vtype=GRB.BINARY, name="e_p")


# alfa_jil : alfa[equipo,equipo,fecha]
# binaria, toma el valor 1 si el equipo j termina con menos puntos
# que el equipo i en el
# MEJOR conjunto de
# resultados futuros para el equipo i considerando que
# se está en la fecha l
alfa_m = m.addVars(I,I,F, vtype=GRB.BINARY, name="alfa_m")

# alfa_jil : alfa[equipo,equipo,fecha]
# binaria, toma el valor 1 si el equipo j tiene termina
# con menos puntos que el equipo i, en el PEOR conjunto de
# resultados futuros para el equipo i considerando que
# se está en la fecha l
alfa_p = m.addVars(I,I,F, vtype=GRB.BINARY, name="alfa_p")

#beta_il: beta[equipo,fecha]
#discreta, indica la mejor posicion
#que puede alcanzar el equipo i al final del 
#torneo, mirando desde la fecha l en el MEJOR
#conjunto de resultados futuros para el equipo i
beta_m = m.addVars(I,F, vtype=GRB.INTEGER, name="beta_m")

#beta_il: beta[equipo, fecha]
#discreta, indica la mejor posicion
#que puede alcanzar el equipo i al final del 
#torneo, mirando desde la fecha l en el PEOR
#conjunto de resultados futuros para el equipo i
beta_p = m.addVars(I,F, vtype=GRB.INTEGER, name="beta_p")


print(f"** VARIABLES TIME: {time.time() - start_model}")


#####################
#*  RESTRICCIONES  *#
#####################

# R2
m.addConstrs((quicksum(x[n, f] for f in F) == 1 for n in N), name="R2")

# R3
m.addConstrs((quicksum(x[n, f] for n in N if EL[i][n] + EV[i][n] == 1) == 1 for i in I
                                                                            for f in F), name="R3")

# R4
for i in I:
  m.addConstr((quicksum(y[i][s] for s in S[i]) == 1), name="R4")


# R6
for i in I:
  m.addConstrs((quicksum(x[n, f] for n in N if EL[i][n] == 1) == quicksum(y[i][s] for s in S[i] if L[s][f] == 1) for f in F), name="R6")

# R7
for i in I:
  m.addConstrs((quicksum(x[n, f] for n in N if EV[i][n] == 1) == quicksum(y[i][s] for s in S[i] if L[s][f] == 0) for f in F), name="R7")

# R8
m.addConstrs((x[n,f] == (v_m[n,i,l,f] + e_m[n,i,l,f] + a_m[n,i,l,f])
                                                                  for n in N
                                                                  for i in I
                                                                  for f in F
                                                                  for l in F
                                                                  if  f > l),name="R8")

# R9
m.addConstrs((x[n,f] == (v_p[n,i,l,f] + e_p[n,i,l,f] + a_p[n,i,l,f])
                                                                  for n in N
                                                                  for i in I
                                                                  for f in F
                                                                  for l in F
                                                                  if  f > l),name="R9")

# R10
m.addConstrs(((p_m[j,i,l,f] == PI[j] + quicksum(quicksum(R[j][n] * x[n,theta] for n in N if EL[j][n] + EV[j][n]  == 1) for theta in F if theta > 5 and theta <= l)
           + quicksum(quicksum(3 * v_m[n,i,l,theta] for theta in F if theta > l and theta <= f) for n in N if EL[j][n] == 1)
           + quicksum(quicksum(3 * a_m[n,i,l,theta] for theta in F if theta > l and theta <= f) for n in N if EV[j][n] == 1)
           + quicksum(quicksum(e_m[n,i,l,theta] for theta in F if theta > l and theta <= f) for n in N if EL[j][n] + EV[j][n] == 1))
                        for j in I 
                        for i in I 
                        for f in F 
                        for l in F), name="R10")

# R11
m.addConstrs(((p_p[j,i,l,f] == PI[j] + quicksum(quicksum(R[j][n] * x[n,theta] for n in N if EL[j][n] + EV[j][n]  == 1) for theta in F if theta > 5 and theta <= l)
           + quicksum(quicksum(3 * v_p[n,i,l,theta] for theta in F if theta > l and theta <= f) for n in N if EL[j][n] == 1)
           + quicksum(quicksum(3 * a_p[n,i,l,theta] for theta in F if theta > l and theta <= f) for n in N if EV[j][n] == 1)
           + quicksum(quicksum(e_p[n,i,l,theta] for theta in F if theta > l and theta <= f) for n in N if EL[j][n] + EV[j][n] == 1))
                        for j in I 
                        for i in I 
                        for f in F 
                        for l in F), name="R11")

#R12
m.addConstrs((((M * (alfa_m[j, i, l]) >= p_m[i, i, l, F[-1]] - p_m[j, i, l, F[-1]])) for i in I
                                                                           for j in I
                                                                           for l in F), name="R12")
# R13
m.addConstrs((((M - M * alfa_m[j,i, l] >= p_m[j, i, l, F[-1]] - p_m[i, i, l, F[-1]])) for i in I
                                                                          for j in I
                                                                          for l in F), name="R13")

# R14
m.addConstrs((((M * (alfa_p[j, i, l]) >= p_p[i, i, l, F[-1]] - p_p[j, i, l, F[-1]])) for i in I
                                                                           for j in I
                                                                           for l in F), name="R14")
#R15
m.addConstrs((((M - M * alfa_p[j, i, l] >= p_p[j, i, l, F[-1]] - p_p[i, i, l, F[-1]])) for i in I
                                                                          for j in I
                                                                          for l in F), name="R15")

#R16
for i in I:
  m.addConstrs(((beta_m[i,l]==len(I)-(quicksum(alfa_m[j,i,l] for j in I if i!=j))) for l in F),name="R16")


# R17
for i in I:
  m.addConstrs(((beta_p[i,l]== 1+(quicksum((1-alfa_p[j,i,l]) for j in I if i!=j))) for l in F),name="R17")






#''' 
#10 y 11 estaban mal escritas.
# R10
#m.addConstrs(((p_m[j,i,l,f] == PI[j] + quicksum(quicksum(R[i][n] * x[n,theta] for n in N if EL[i][n] + EV[i][n]  == 1) for theta in F if theta > 5 and theta <= l)
#           + quicksum(quicksum(3 * v_m[n,i,theta,fpython] for theta in F if theta > l and theta <= f) for n in N if EL[j][n] == 1) 
#           + quicksum(quicksum(3 * a_m[n,i,theta,f] for theta in F if theta > l and theta <= f) for n in N if EV[j][n] == 1) 
#           + quicksum(quicksum(e_m[n,i,theta,f] for theta in F if theta > l and theta <= f) for n in N if EL[j][n] + EV[j][n] == 1)) 
#                        for j in I 
#                        for i in I 
#                        for f in F 
#                        for l in F), name="R10")

# R11
#m.addConstrs(((p_p[j,i,l,f] == PI[j] + quicksum(quicksum(R[i][n] * x[n,theta] for n in N if EL[i][n] + EV[i][n]  == 1) for theta in F if theta > 5 and theta <= l)
#           + quicksum(quicksum(3 * v_p[n,i,theta,f] for theta in F if theta > l and theta <= f) for n in N if EL[j][n] == 1) 
#           + quicksum(quicksum(3 * a_p[n,i,theta,f] for theta in F if theta > l and theta <= f) for n in N if EV[j][n] == 1) 
#           + quicksum(quicksum(e_p[n,i,theta,f] for theta in F if theta > l and theta <= f) for n in N if EL[j][n] + EV[j][n] == 1)) 
#                        for j in I 
#                        for i in I 
#                        for f in F 
#                        for l in F), name="R11")

#R12
#m.addConstrs((((M - M * alfa_m[i, j, l] >= p_m[j, i, l, F[-1]] - p_m[i, i, l, F[-1]])) for i in I
#                                                                          for j in I
#                                                                          for l in F), name="R12")

# R13                                                                                        
#m.addConstrs((((M * (alfa_p[i, j, l]) >= p_p[i, i, l, F[-1]] - p_p[j, i, l, F[-1]])) for i in I
#                                                                           for j in I
#                                                                           for l in F), name="R13")

#R14
#for i in I:
#  m.addConstrs(((beta_m[i,l]==len(I)-(quicksum(alfa_m[i,j,l] for j in I if i!=j))) for l in F),name="R14")


# R15
#for i in I:
#  m.addConstrs(((beta_p[i,l]==len(I)-(quicksum(alfa_p[i,j,l] for j in I if i!=j))) for l in F),name="R15")

#'''
print(f"** RESTRICTIONS TIME: {time.time() - start_model}")
########################
#*  FUNCION OBJETIVO  *#
########################

m.setObjective(quicksum(quicksum(beta_p[i,l]-beta_m[i,l] for i in I) for l in F), GRB.MAXIMIZE)
### FOs de prueba
#m.setObjective(quicksum(quicksum(x[n, f] for n in N) for f in F), GRB.MAXIMIZE)
#m.setObjective(quicksum(p_m[i,i,6,10] for i in I), GRB.MAXIMIZE)




m.optimize()

print(f"** TOTAL TIME: {time.time() - START_TIME}")

ModelStats.parse_gurobi_output(m.getVars(), stats.matches, S_full)
ModelStats.check_valid_output()



wb = opx.Workbook()
wb.save("salida.xlsx")
def escribir_hoja(hoja,equipo, fechas, fecha_inicio, equipos, partidos_por_fecha):
    hoja = wb[hoja]
    hoja.title = equipo
    fecha_i= fecha_inicio
    for fecha in range(fechas):
        row = 1+(fecha)*(len(equipos)+partidos_por_fecha+1)
        column = 3
        fecha_i = fecha_inicio
        for fecha in range(fechas):
            hoja.cell(row=row, column=column, value ="fecha"+ str(fecha_i))
            column+=2
            fecha_i +=1

    for fecha in range(fechas): 
        row = 1+(fecha)*(len(equipos)+partidos_por_fecha+1)
        column = 1
        hoja.cell(row=row, column=column, value ="Puntaje Conocido")
        hoja.cell(row = row, column = 1+len(equipos)*2, value="Alfa")
        hoja.cell(row = row, column = 3+len(equipos)*2, value="Beta")

    for i in range(fechas):
        row = (len(equipos)+partidos_por_fecha+1)*i + 2+partidos_por_fecha
        column = 1
        for j in range(fechas+1):
            for equipo in equipos:
                hoja.cell(row=row, column=column, value=equipo)
                row+=1
            column +=2
            row = (len(equipos)+partidos_por_fecha+1)*i+2+partidos_por_fecha
    for i in range(fechas):
        row = 2 + i*(len(equipos)+partidos_por_fecha+1)
        column = 3
        fecha_i = fecha_inicio
        for j in range(fechas):
            for n in range(len(dicc_n_f[fecha_i])):
              hoja.cell(row=n+row, column=column, value=stats.matches[dicc_n_f[fecha_i][n]]['home'] + "-" + stats.matches[dicc_n_f[fecha_i][n]]['away'])
            column +=2
            fecha_i+=1
    stop = 0
    for i in range(fechas):
        row = 2 + i*(len(equipos)+partidos_por_fecha+1)
        column = 4
        fecha_i = fecha_inicio
        for j in range(fechas):
            for n in range(len(dicc_n_f[fecha_i])):
              hoja.cell(row=n+row, column= column, value=stats.matches[dicc_n_f[fecha_i][n]]['winner'])
            column +=2
            fecha_i+=1
            if j == stop:
              stop += 1
              break
    for i in range(1, fechas):
        row = 2 + (i - 1)*(len(equipos)+partidos_por_fecha+1)
        column = 4 + 2 * i
        fecha_i = fecha_inicio + i - 1
        for j in range(i + 1, fechas + 1):
            fecha_j = fecha_inicio + j - 1
            for n in range(len(dicc_n_f[fecha_j])):
              if "ME" in hoja.title:
                if (dicc_n_f[fecha_j][n], hoja.title[2:], fecha_i, fecha_j) in v_m_list:
                  hoja.cell(row=n+row, column= column, value='H')
                elif (dicc_n_f[fecha_j][n], hoja.title[2:], fecha_i, fecha_j) in a_m_list:
                  hoja.cell(row=n+row, column= column, value='A')
                elif (dicc_n_f[fecha_j][n], hoja.title[2:], fecha_i, fecha_j) in e_m_list:
                  hoja.cell(row=n+row, column= column, value='D')                  
              else:
                if (dicc_n_f[fecha_j][n], hoja.title[2:], fecha_i, fecha_j) in v_p_list:
                  hoja.cell(row=n+row, column= column, value='H')
                elif (dicc_n_f[fecha_j][n], hoja.title[2:], fecha_i, fecha_j) in a_p_list:
                  hoja.cell(row=n+row, column= column, value='A')
                elif (dicc_n_f[fecha_j][n], hoja.title[2:], fecha_i, fecha_j) in e_p_list:
                  hoja.cell(row=n+row, column= column, value='D')
            column +=2
    for i in range(fechas):
        row = (len(equipos)+partidos_por_fecha+1)*i + 2+partidos_por_fecha
        column = 2
        for j in range(fechas+1):
            for equipo in equipos:
                hoja.cell(row=row, column=column, value=PI[equipo])
                row+=1
            row = (len(equipos)+partidos_por_fecha+1)*i+2+partidos_por_fecha
    stop=0
    for i in range(fechas):
        row = 2+partidos_por_fecha + i*(len(equipos)+partidos_por_fecha+1)
        column = 4
        fecha_i = fecha_inicio
        for j in range(fechas):
            c = 0
            for equipo in equipos:
              if "ME" in hoja.title:
                hoja.cell(row=c + row, column=column, value=p_m_dic[(equipo, hoja.title[2:], fecha_inicio + j, fecha_inicio + j)])
                c+=1
              else:
                hoja.cell(row=c + row, column=column, value=p_p_dic[(equipo, hoja.title[2:], fecha_inicio + j, fecha_inicio + j)])
                c+=1
            column +=2
            fecha_i+=1
            if j == stop:
              stop += 1
              break
    for i in range(0, fechas):
        row = (len(equipos)+partidos_por_fecha+1)*i + 2+partidos_por_fecha
        column = 6 + 2 * i
        fecha_i = fecha_inicio + i -1
        for j in range(i + 1, fechas):
            fecha_j = fecha_inicio + j -1
            c = 0
            for equipo in equipos:
              if "ME" in hoja.title:
                  hoja.cell(row=c+row, column= column, value=p_m_dic[(equipo, hoja.title[2:], fecha_inicio + i, fecha_inicio + j)])
                  c+=1
              else:
                  hoja.cell(row=c+row, column= column, value=p_p_dic[(equipo, hoja.title[2:], fecha_inicio + i, fecha_inicio + j)])
                  c+=1
            column +=2
    for i in range(fechas):
        column = 13
        for j in range(1, fechas + 1):
            row = (len(equipos)+partidos_por_fecha+1)*(j - 1) + 2+partidos_por_fecha
            c = 0
            for equipo in equipos:
              if "ME" in hoja.title:
                #print(str((equipo, hoja.title[2:], fecha_inicio + j - 1)) + " ALFA_M = " + str(alfa_m_dic[(equipo, hoja.title[2:], fecha_inicio + j - 1)]))
                #valor = str(alfa_m_dic[(equipo, hoja.title[2:], fecha_inicio + j - 1)])
                hoja.cell(row=c+row, column=column, value=str(alfa_m_dic[(equipo, hoja.title[2:], fecha_inicio + j - 1)]))
                c += 1
    for i in range(fechas):
        column = 13
        for j in range(1, fechas + 1):
            row = (len(equipos)+partidos_por_fecha+1)*(j - 1) + 2+partidos_por_fecha
            c = 0
            for equipo in equipos:
              if "PE" in hoja.title:
                #print(str((equipo, hoja.title[2:], fecha_inicio + j - 1)) + " ALFA_P = " + str(alfa_p_dic[(equipo, hoja.title[2:], fecha_inicio + j - 1)]))
                #valor = str(alfa_p_dic[(equipo, hoja.title[2:], fecha_inicio + j - 1)])
                hoja.cell(row=c+row, column=column, value=str(alfa_p_dic[(equipo, hoja.title[2:], fecha_inicio + j - 1)]))
                c+=1
    for i in range(fechas):
        row = (len(equipos)+partidos_por_fecha+1)*i + 2+partidos_por_fecha
        column = (fechas+2)*2+1
        for j in range(fechas+1):
            for equipo in equipos:
                if "ME" in hoja.title and equipo in hoja.title[2:]:
                    hoja.cell(row=row, column=column, value=beta_m_dic[(equipo, fecha_inicio + i)])
                    row+=1
                elif "PE" in hoja.title and equipo in hoja.title[2:]: 
                    hoja.cell(row=row, column=column, value=beta_p_dic[(equipo, fecha_inicio + i)])
                    row+=1
                else:
                    hoja.cell(row=row, column=column, value="-")
                    row+=1
            row = (len(equipos)+partidos_por_fecha+1)*i+2+partidos_por_fecha
    wb.save("salida.xlsx")


dicc_n_f = {}
v_m_list = []
v_p_list = []

a_m_list = []
a_p_list = []

e_m_list = []
e_p_list = []

p_m_dic = {}
p_p_dic = {}

alfa_m_dic = {}
alfa_p_dic = {}

beta_m_dic = {}
beta_p_dic = {}

for i in range(6, 11):
  dicc_n_f[i] = list()
for var in m.getVars():
  if "value 1" in str(var) and "x" in str(var):
    i = str(var).index("[")
    f = str(var).index("]")
    numeros = str(var)[i+1:f]
    N,f = numeros.split(",")
    dicc_n_f[int(f)].append(int(N))
  elif "value 1" in str(var) and "v_m" in str(var):
    i = str(var).index("[")
    f = str(var).index("]")
    numeros = str(var)[i+1:f]
    n, i, l, f = numeros.split(",")
    v_m_list.append((int(n), i, int(l), int(f)))
  elif "value 1" in str(var) and "v_p" in str(var):
    i = str(var).index("[")
    f = str(var).index("]")
    numeros = str(var)[i+1:f]
    n, i, l, f = numeros.split(",")
    v_p_list.append((int(n), i, int(l), int(f)))
  elif "value 1" in str(var) and "a_m" in str(var) and not("alfa_m" in str(var)) and not("beta_m" in str(var)):
    i = str(var).index("[")
    f = str(var).index("]")
    numeros = str(var)[i+1:f]
    n, i, l, f = numeros.split(",")
    a_m_list.append((int(n), i, int(l), int(f)))
  elif "value 1" in str(var) and "a_p" in str(var) and not("alfa_p" in str(var)) and not("beta_p" in str(var)):
    i = str(var).index("[")
    f = str(var).index("]")
    numeros = str(var)[i+1:f]
    n, i, l, f = numeros.split(",")
    a_p_list.append((int(n), i, int(l), int(f)))
  elif "value 1" in str(var) and "e_m" in str(var):
    i = str(var).index("[")
    f = str(var).index("]")
    numeros = str(var)[i+1:f]
    n, i, l, f = numeros.split(",")
    e_m_list.append((int(n), i, int(l), int(f)))
  elif "value 1" in str(var) and "e_p" in str(var):
    i = str(var).index("[")
    f = str(var).index("]")
    numeros = str(var)[i+1:f]
    n, i, l, f = numeros.split(",")
    e_p_list.append((int(n), i, int(l), int(f)))
  elif "p_m" in str(var):
    i = str(var).index("[")
    f = str(var).index("]")
    i_value = str(var).index("(") + 7
    j_value = str(var).index(")") - 2
    numeros = str(var)[i+1:f]
    value = str(var)[i_value:j_value]
    j, i, l, f = numeros.split(",")
    p_m_dic[(j, i, int(l), int(f))] = value
  elif "p_p" in str(var):
    i = str(var).index("[")
    f = str(var).index("]")
    i_value = str(var).index("(") + 7
    j_value = str(var).index(")") - 2
    numeros = str(var)[i+1:f]
    value = str(var)[i_value:j_value]
    j, i, l, f = numeros.split(",")
    p_p_dic[(j, i, int(l), int(f))] = value
  elif "alfa_m" in str(var):
    i = str(var).index("[")
    f = str(var).index("]")
    i_value = str(var).index("(") + 7
    j_value = str(var).index(")") - 2
    numeros = str(var)[i+1:f]
    value = str(var)[i_value:j_value]
    j, i, l = numeros.split(",")
    alfa_m_dic[(j, i, int(l))] = value
  elif "alfa_p" in str(var):
    i = str(var).index("[")
    f = str(var).index("]")
    i_value = str(var).index("(") + 7
    j_value = str(var).index(")") - 2
    numeros = str(var)[i+1:f]
    value = str(var)[i_value:j_value]
    j, i, l = numeros.split(",")
    alfa_p_dic[(j, i, int(l))] = value
  elif "beta_m" in str(var):
    i = str(var).index("[")
    f = str(var).index("]")
    i_value = str(var).index("(") + 7
    j_value = str(var).index(")") - 2
    numeros = str(var)[i+1:f]
    value = str(var)[i_value:j_value]
    i, l = numeros.split(",")
    beta_m_dic[(i, int(l))] = value
  elif "beta_p" in str(var):
    i = str(var).index("[")
    f = str(var).index("]")
    i_value = str(var).index("(") + 7
    j_value = str(var).index(")") - 2
    numeros = str(var)[i+1:f]
    value = str(var)[i_value:j_value]
    i, l = numeros.split(",")
    beta_p_dic[(i, int(l))] = value

hojas = []
for equipo in stats.teams.keys():
    mejor = "ME" + equipo
    peor = "PE" + equipo
    hojas.append(mejor)
    hojas.append(peor)
for hoja in hojas:
    wb.create_sheet(hoja)
    escribir_hoja(hoja,hoja, 5, 6, stats.teams.keys(), 3)


#for i in beta_m_dic:
#  print(str(i) + " MEJOR LUGAR ES " + str(beta_m_dic[i]) + " PEOR LUGAR ES " + str(beta_p_dic[i]))